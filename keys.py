import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import csv
import time

# Function to update a text-based progress bar during processing
def update_progress_bar(total, progress):
    bar_length = 40  # Set the length of the progress bar
    if total > 0:
        fraction_completed = progress / total
    else:
        fraction_completed = 0
    arrow = int(fraction_completed * bar_length - 1) * "=" + ">"
    padding = (bar_length - len(arrow)) * ' '
    print(f"\rProgress: [{arrow}{padding}] {progress}/{total} URLs processed", end='', flush=True)

# Function to color text in terminal output
def colored(text, color):
    colors = {
        "red": "\033[31m",
        "green": "\033[32m",
        "yellow": "\033[33m",
        "blue": "\033[34m",
        "magenta": "\033[35m",
        "cyan": "\033[36m",
        "white": "\033[37m",
    }
    reset = "\033[0m"  # Resets the color to default
    return f"{colors.get(color, reset)}{text}{reset}"

# Function to convert a keyword into a hyphenated format
def convert_to_hyphenated(keyword):
    # Convert keyword to string in case it's not
    keyword_str = str(keyword)
    return keyword_str.replace(' ', '-').lower()

# Function to check if the keyword is present in a paragraph tag
def keyword_in_paragraph(keyword, driver):
    paragraphs = driver.find_elements(By.TAG_NAME, 'p')
    for paragraph in paragraphs:
        if keyword.lower() in paragraph.text.lower():
            return 'Pass'
    return 'Fail'

# Function to check if the URL is a document (PDF, DOC, DOCX)
def is_document(url):
    return re.search(r'\.(pdf|docx?)(\?.*)?$', url.lower()) is not None


# Function to find the next available file name with an iterative suffix
def next_available_filename(base_name, extension):
    counter = 1
    while True:
        new_name = f"{base_name}-{counter}{extension}"
        if not os.path.exists(new_name):
            return new_name
        counter += 1

# File paths for input and output
input_file_path = 'rank.xlsx'
output_file_path = 'output.csv'

# Base name and extension for output file
base_output_file_name = 'output'
output_file_extension = '.csv'

# Find the next available file name
output_file_path = next_available_filename(base_output_file_name, output_file_extension)

# Selenium WebDriver configuration for headless Chrome
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")

service = Service(executable_path='/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=service, options=chrome_options)

# Initialize a list to store results
results = []

# Load the Excel workbook and select the active sheet
wb = load_workbook(filename=input_file_path)
ws = wb.active

# Count the number of URLs to process
total_urls = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value and row[1].value)
processed_urls = 0

# Process each URL in the Excel file
for row in ws.iter_rows(min_row=2, values_only=True):
    keyword, url = row[0], row[1]

    if not (keyword and url):
        continue  # Skip empty rows

    # Convert keyword to hyphenated format
    hyphenated_keyword = convert_to_hyphenated(keyword)

    try:
        # Open the URL with Selenium WebDriver
        driver.get(url)
        time.sleep(2)  # Wait for the page to load

        # Check if the keyword is in the page title
        page_title = driver.title
        keyword_in_title = 'Pass' if keyword.lower() in page_title.lower() else 'Fail'

        # Check if the keyword is in the H1 tag of the page
        try:
            h1_tag = driver.find_element(By.TAG_NAME, 'h1').text
            keyword_in_h1 = 'Pass' if keyword.lower() in h1_tag.lower() else 'Fail'
        except:
            keyword_in_h1 = 'Fail'  # Fail if H1 tag is not found or other errors

        # Check if the hyphenated keyword is in the URL
        keyword_in_url = 'Pass' if hyphenated_keyword in url.lower() else 'Fail'

        # Check if the keyword is in at least one paragraph tag
        keyword_in_paragraph_tag = keyword_in_paragraph(keyword, driver)

        # Check if the hyphenated keyword is in image URLs and alt attributes
        images = driver.find_elements(By.TAG_NAME, 'img')
        keyword_in_image_url = 'Fail'
        keyword_in_alt_attribute = 'Fail'
        for image in images:
            if hyphenated_keyword in image.get_attribute('src').lower():
                keyword_in_image_url = 'Pass'
            if hyphenated_keyword in (image.get_attribute('alt') or '').lower():
                keyword_in_alt_attribute = 'Pass'

        results.append((url, keyword, keyword_in_title, keyword_in_h1, keyword_in_url, keyword_in_paragraph_tag, keyword_in_image_url, keyword_in_alt_attribute))

    except Exception as e:
        print(f"Error processing URL {url}: {e}")
        continue

    processed_urls += 1
    update_progress_bar(total_urls, processed_urls)  # Update the progress bar

# Print the final message after processing
output_file_full_path = os.path.join(os.getcwd(), output_file_path)
completion_message = colored("Processing Complete! ", "magenta")
path_message = colored(f"Your document is saved at {output_file_full_path}", "white")
print('\n' + completion_message + path_message)

# Write results to the new CSV file
with open(output_file_path, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['URL', 'Keyword', 'Keyword in Title', 'Keyword In H1', 'Keyword in URL', 'Keyword in Paragraph', 'Keyword in Image URL', 'Keyword in Alt Attribute'])
    for row in results:
        writer.writerow(row)

# Quit the WebDriver
driver.quit()
