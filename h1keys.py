import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import csv
import time

# Function to update a text-based progress bar during processing
def update_progress_bar(total, progress):
    bar_length = 40  # Set the length of the progress bar
    if total > 0:
        fraction_completed = progress / total
    else:
        fraction_completed = 0
    arrow = int(fraction_completed * bar_length - 1) * "=" + ">"
    padding = (bar_length - len(arrow)) * ' '
    print(f"\rProgress: [{arrow}{padding}] {progress}/{total} URLs processed", end='', flush=True)

# Function to color text in terminal output
def colored(text, color):
    colors = {
        "red": "\033[31m",
        "green": "\033[32m",
        "yellow": "\033[33m",
        "blue": "\033[34m",
        "magenta": "\033[35m",
        "cyan": "\033[36m",
        "white": "\033[37m",
    }
    reset = "\033[0m"  # Resets the color to default
    return f"{colors.get(color, reset)}{text}{reset}"

# File paths for input and output
input_file_path = 'h1keywords.xlsx'
output_file_path = 'output.csv'

# Selenium WebDriver configuration for headless Chrome
chrome_options = Options()
chrome_options.add_argument("--headless=new")
chrome_options.add_argument("--disable-gpu")

service = Service(executable_path='/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=service, options=chrome_options)

# Initialize a list to store results
results = []

# Load the Excel workbook and select the active sheet
wb = load_workbook(filename=input_file_path)
ws = wb.active

# Count the number of URLs to process
total_urls = sum(1 for row in ws.iter_rows(min_row=2) if any(cell.value for cell in row))
processed_urls = 0

# Determine the column indices for 'Keyword' and 'Current URL'
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
keyword_column = header.index('Keyword')
url_column = header.index('Current URL')

# Process each URL in the Excel file
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    if not any(cell for cell in row):
        continue  # Skip empty rows

    keyword = row[keyword_column]
    url = row[url_column]

    driver.get(url)  # Open the URL with Selenium WebDriver
    time.sleep(2)  # Wait for the page to load

    # Check if the keyword is in the H1 tag of the page
    try:
        h1_tag = driver.find_element(By.TAG_NAME, 'h1').text
        pass_fail = 'Pass' if keyword.lower() in h1_tag.lower() else 'Fail'
    except:
        pass_fail = 'Fail'  # Fail if H1 tag is not found or other errors

    results.append((url, keyword, pass_fail))  # Append result to list

    processed_urls += 1
    update_progress_bar(total_urls, processed_urls)  # Update the progress bar

# Print the final message after processing
output_file_full_path = os.path.join(os.getcwd(), output_file_path)
completion_message = colored("Processing Complete! ", "magenta")
path_message = colored(f"Your document is saved at {output_file_full_path}", "white")
print('\n' + completion_message + path_message)

# Group results by URL
grouped_results = {}
for url, keyword, pass_fail in results:
    if url not in grouped_results:
        grouped_results[url] = []
    grouped_results[url].append((keyword, pass_fail))

# Write results to a CSV file
with open(output_file_path, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['URL', 'Keyword', 'Pass/Fail'])
    for url, checks in grouped_results.items():
        for keyword, pass_fail in checks:
            writer.writerow([url, keyword, pass_fail])

# Quit the WebDriver
driver.quit()
